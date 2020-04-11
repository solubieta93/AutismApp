import openpyxl
import collections
import datetime
from datetime import timedelta
import os
from Database.manager import Manager
from datetime import date


def read_data(username, address):
    print address
    list_excel = os.listdir(address)
    for excel in list_excel:
        print excel
        excel_document = openpyxl.load_workbook(os.path.join(address, excel))
        sheets = excel_document.get_sheet_names()

        m = Manager()
        current_medicaid_id = ''
        current_key = ''
        for i in range(0, len(sheets)):

            if sheets[i] == 'Programs':
                table_name = 'programs'

            elif sheets[i] == 'Maladaptive Behaviors':
                table_name = 'behaviors'

            else:
                table_name = sheets[i]

            sheet = excel_document.get_sheet_by_name(sheets[i])

            if table_name == 'Name':
                multiple_cells = sheet['A1':'E1']
                for row in multiple_cells:
                    list_data = []
                    for cell in row:
                        list_data.append(cell.value)
                    current_key = str(list_data[3])
                    current_medicaid_id = str(list_data[3])

                    analyst = list_data[4] if list_data[4] is not None else ''
                    if not m.existing_child(username, list_data[0], list_data[1], list_data[2], str(list_data[3]), analyst)[
                        0]:
                        m.add_kid(username, list_data[0], list_data[1], list_data[2], str(list_data[3]), analyst)
                continue

            elif table_name == 'programs' or table_name == 'behaviors':

                list_data = collections.OrderedDict()

                cell_row = -1
                current_beh_or_program = ''
                dict_aux = collections.OrderedDict()
                dict_bl_aux = collections.OrderedDict()


                list_data[current_key] = collections.OrderedDict()

                # all_rows = sheet.rows
                # for row_i in all_rows[:]:
                # PARA TRABAJAR EN MI LAPTOP
                for row_i in sheet.rows:
                    count_columns = 0
                    count_none = 0
                    bl_founded = False
                    count_bl_added = 0
                    for cell_i in row_i:
                        if cell_i.value is None:
                            count_none += 1
                        if count_none >= 3:
                            break
                        count_columns += 1

                        # Clear dictionary because was found another one
                        if str(cell_i.value) == 'Maladaptive Behaviors' or str(cell_i.value) == 'Programs':
                            cell_row = cell_i.row
                            dict_aux.clear()
                            dict_bl_aux.clear()

                        # create keys on the dict if I keep in the date's row
                        elif cell_row is cell_i.row and cell_i.value is not None:
                            if cell_i.value == 'BL':
                                bl_founded = True
                            else:
                                if bl_founded:
                                    dict_aux[cell_i.value] = -1
                                else:
                                    dict_bl_aux[cell_i.value] = -1

                        # create key --> name of behavior
                        elif cell_i.column == 'A':
                            list_data[current_key][cell_i.value] = dict_aux.copy()
                            current_beh_or_program = cell_i.value

                        # save data as value of behavior name key on the dict
                        elif cell_i.value is not None:
                            if cell_i.value == 'BL':
                                bl_founded = True
                                continue
                            # key = sheet.cell(row=cell_row - 1, column=count_columns-1).value
                            key = sheet.cell(row=cell_row, column=cell_i.col_idx).value
                            if bl_founded:
                                list_data[current_key][current_beh_or_program][key] = cell_i.value
                            else:
                                if count_bl_added > 0:
                                    continue
                                exist = True if current_beh_or_program in m.list_behavior_or_program(table_name,
                                                                                                     str(
                                                                                                         current_key)) \
                                    else False
                                if not exist:
                                    measure = 'Time' if table_name == 'behaviors' else '% of Opportunities'
                                    m.add(table_name, current_key, current_beh_or_program, measure)

                                baseline = m.list_baseline(table_name, str(current_key), str(current_beh_or_program))
                                if not baseline:
                                    exist_end_date = True
                                    if len(dict_bl_aux.keys()) > 1:
                                        date1 = dict_bl_aux.keys()[0]
                                        date2 = dict_bl_aux.keys()[1]
                                    else:
                                        date1 = dict_bl_aux.keys()[0]
                                        try:
                                            n_date = date1.split('-')
                                        except:
                                            n_date = str(date1.date()).split('-')

                                        n_date = date(int(n_date[0]), int(n_date[1]), int(n_date[2]))
                                        date2 = str(n_date + timedelta(days=3))
                                        exist_end_date = False

                                    m.add_baseline(table_name, str(current_key), str(current_beh_or_program),
                                                   cell_i.value,
                                                   'Weeks', date1, date2, exist_end_date)
                                    count_bl_added += 1

                # self.table_name, self.medicaid_id, self.name_program_or_behavior, date
                for medicaid_id in list_data.keys():
                    for beh_or_program in list_data[medicaid_id].keys():
                        for date_i in list_data[medicaid_id][beh_or_program]:
                            if date_i is not None:

                                if type(date_i) is datetime.datetime:
                                    n_date = str(date_i.date())
                                else:
                                    n_date = date_i.split('-')
                                    n_date = str(date(int(n_date[0]), int(n_date[1]), int(n_date[2])))

                                if not m.existing_data(table_name, str(medicaid_id), str(beh_or_program), n_date):
                                    value = list_data[medicaid_id][beh_or_program][date_i]
                                    try:
                                        value = int(value)
                                        m.add_data_kid(table_name, str(medicaid_id), str(beh_or_program), n_date, value)
                                    except ValueError:
                                        m.add_data_kid(table_name, str(medicaid_id), str(beh_or_program), n_date, -1, value)

            else:
                sheet_name = table_name
                table_name = sheet_name.split('_')[1]
                part_table_name = sheet_name.split('_')[0]
                sheet = excel_document.get_sheet_by_name(sheet_name)
                if part_table_name == 'lto':
                    # range_string = u'E{}'.format(str(sheet.get_highest_row()))
                    range_string = u'E{}'.format(str(sheet.max_row))
                    multiple_cells = sheet['A2':range_string]
                    for row in multiple_cells:
                        list_data = []
                        for cell_i in row:
                            if cell_i.value is None:
                                break
                            list_data.append(cell_i.value)
                        if list_data:
                            # print str(list_data[0])

                            b, s, data = m.lto_data(table_name, current_medicaid_id, str(list_data[0]))
                            # print s, b, data

                            if not b:
                                m.add_lto(table_name, current_medicaid_id, str(list_data[0]), list_data[1], list_data[2],
                                          str(list_data[3]))
                                # else:
                                #     m.update_components_table(sheets[i], ['lto_count', 'lto_time_count', 'lto_time'], [list_data[2],
                                #                                                                                         list_data[3],
                                #                                                                                         str(list_data[4])],
                                #                               str(list_data[0]), str(list_data[1]))

                elif part_table_name == 'sto':
                    # range_string = u'F{}'.format(str(sheet.get_highest_row()))
                    range_string = u'F{}'.format(str(sheet.max_row))
                    multiple_cells = sheet['A2':range_string]
                    current_sto = ''
                    count_sto_number = 0
                    for row in multiple_cells:
                        list_data = []
                        for cell_i in row:
                            if cell_i.value is None:
                                break
                            list_data.append(cell_i.value)
                        if list_data:
                            if not list_data[0] == current_sto:
                                current_sto = list_data[0]
                                count_sto_number = 1
                            else:
                                count_sto_number += 1
                            count_sto_string = 'STO ' + str(count_sto_number) + '--'
                            # print str(list_data)
                            if sheet_name == 'sto_programs' and int(list_data[2]) is -1:
                                string_rep = count_sto_string + ' No less than ' + str(list_data[1]) + \
                                             ' incidents per week for ' + str(list_data[3]) + \
                                             ' consecutive ' + str(list_data[4])
                            elif sheet_name == 'sto_behaviors' and int(list_data[1]) is -1:
                                string_rep = count_sto_string + ' No more than ' + str(
                                    list_data[2]) + ' incidents per week for ' \
                                             + str(list_data[3]) + ' consecutive ' + str(list_data[4])
                            else:
                                string_rep = count_sto_string + ' Between ' + str(list_data[1]) + ' and ' + str(
                                    list_data[2]) + ' incidents per week for ' \
                                             + str(list_data[3]) + ' consecutive ' + str(list_data[4])

                            exist = m.sto_data_string(table_name, current_medicaid_id, str(list_data[0]))
                            # if str(list_data[0]) == '9543545952':
                            #     print str(list_data[1])
                            #     print 'exist'
                            #     print exist
                            if exist[0]:
                                exist = exist[2]
                                exist = [item_str[3] for item_str in exist]

                                exist1 = map(lambda x: x.split('(MASTERED')[0], exist)

                                if string_rep not in exist1:
                                    m.add_sto(table_name, current_medicaid_id, str(list_data[0]), list_data[1],
                                              list_data[2],
                                              list_data[3], str(list_data[4]), string_rep)
                            else:
                                m.add_sto(table_name, current_medicaid_id, str(list_data[0]), list_data[1], list_data[2],
                                          list_data[3], str(list_data[4]))


if __name__ == '__main__':
    # read_data('yessica', 'E:\AutimsApp\Excels')
    read_data('yessica', 'D:\proyects\AutismApp\Excels')
